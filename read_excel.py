import warnings
warnings.filterwarnings('ignore')
import openpyxl

wb = openpyxl.load_workbook('MATRIZ TRATAMIENTO ORAL 2025 170625.xlsx', read_only=True, data_only=True)
print('HOJAS:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n\n========== HOJA: [{sheet_name}] ==========')
    row_count = 0
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True)):
        non_none = [str(x) for x in row if x is not None]
        if non_none:
            print(f'  R{i+1}: {" | ".join(non_none[:20])}')
        row_count = i
    print(f'  (mostrando primeras 8 filas de ~{row_count} totales)')
